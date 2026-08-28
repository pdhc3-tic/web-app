import csv
from datetime import date, timedelta
from io import BytesIO, StringIO
from time import monotonic

import pytest
from django.core.cache import cache
from openpyxl import load_workbook

from apps.sgp.services.workplan_export import EXPORT_COLUMNS, workplan_export_rows
from apps.sgp.tasks import export_to_power_bi
from apps.sgp.tests.factories import (
    ActivityFactory,
    WorkPlanAcaoFactory,
    WorkPlanMetaFactory,
)


pytestmark = pytest.mark.django_db

EXPORT_URL = "/api/v1/sgp/plano-trabalho/exportar/"
POWER_BI_URL = "/api/v1/sgp/plano-trabalho/powerbi/"


def create_action(meta, numero, municipio, *, descricao="Ação exportada"):
    action = WorkPlanAcaoFactory(
        meta=meta,
        numero=numero,
        descricao=descricao,
        quantidade_planejada="2.00",
        valor_unitario="50.00",
        data_inicio=date.today() - timedelta(days=10),
        data_fim=date.today() + timedelta(days=10),
    )
    ActivityFactory(acao=action, municipio=municipio, status="concluido")
    return action


class TestWorkPlanExport:
    def test_csv_has_expected_headers_and_rows(self, auth_client, municipio):
        meta = WorkPlanMetaFactory(numero=1, titulo="Meta de exportação")
        action = create_action(meta, "1.1", municipio)

        response = auth_client.get(f"{EXPORT_URL}?formato=csv")

        assert response.status_code == 200
        assert response["Content-Type"].startswith("text/csv")
        assert response["Content-Disposition"].endswith(".csv\"")
        content = response.content.decode("utf-8-sig")
        rows = list(csv.reader(StringIO(content)))
        assert rows[0] == [label for _, label in EXPORT_COLUMNS]
        assert rows[1][0] == "1 - Meta de exportação"
        assert rows[1][1] == f"{action.numero} - {action.descricao}"
        assert rows[1][6] == "1"

    def test_xlsx_opens_and_has_same_columns(self, auth_client, municipio):
        meta = WorkPlanMetaFactory(numero=1)
        action = create_action(meta, "1.1", municipio)

        response = auth_client.get(f"{EXPORT_URL}?formato=xlsx")

        assert response.status_code == 200
        assert response["Content-Disposition"].endswith(".xlsx\"")
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        assert list(rows[0]) == [label for _, label in EXPORT_COLUMNS]
        assert rows[1][1] == f"{action.numero} - {action.descricao}"

    def test_rejects_unsupported_format(self, auth_client):
        response = auth_client.get(f"{EXPORT_URL}?formato=pdf")

        assert response.status_code == 400
        assert "formato" in response.data

    def test_adt_cannot_export_other_territory_data(
        self, auth_client_adt_rn, municipio_rn, municipio_ce
    ):
        meta = WorkPlanMetaFactory(numero=1)
        own_action = create_action(meta, "1.1", municipio_rn, descricao="Território RN")
        create_action(meta, "1.2", municipio_ce, descricao="Território CE")

        response = auth_client_adt_rn.get(f"{EXPORT_URL}?formato=csv")

        content = response.content.decode("utf-8-sig")
        assert response.status_code == 200
        assert own_action.descricao in content
        assert "Território CE" not in content

    def test_exports_simulated_twelve_month_dataset_under_sixty_seconds(
        self, usuario, municipio
    ):
        meta = WorkPlanMetaFactory(numero=1)
        for index in range(1, 121):
            action = WorkPlanAcaoFactory(
                meta=meta,
                numero=f"1.{index}",
                data_inicio=date.today() - timedelta(days=180),
                data_fim=date.today() + timedelta(days=180),
            )
            ActivityFactory(acao=action, municipio=municipio, status="concluido")

        started_at = monotonic()
        rows = workplan_export_rows(user=usuario)

        assert len(rows) == 120
        assert monotonic() - started_at < 60


class TestPowerBIExport:
    def test_missing_service_token_returns_401(self, api_client):
        response = api_client.get(POWER_BI_URL)

        assert response.status_code == 401

    def test_invalid_service_token_returns_401(self, api_client, settings):
        settings.POWER_BI_SERVICE_TOKEN = "power-bi-test-token"

        response = api_client.get(
            POWER_BI_URL,
            HTTP_AUTHORIZATION="Token invalid-token",
        )

        assert response.status_code == 401

    def test_valid_service_token_returns_latest_snapshot(
        self, api_client, settings, municipio
    ):
        settings.POWER_BI_SERVICE_TOKEN = "power-bi-test-token"
        cache.clear()
        meta = WorkPlanMetaFactory(numero=1)
        action = create_action(meta, "1.1", municipio)
        snapshot = export_to_power_bi()

        response = api_client.get(
            POWER_BI_URL,
            HTTP_AUTHORIZATION="Token power-bi-test-token",
        )

        assert response.status_code == 200
        assert response.data["atualizado_em"] == snapshot["atualizado_em"]
        assert response.data["resultados"][0]["acao"] == (
            f"{action.numero} - {action.descricao}"
        )

    def test_manual_task_replaces_snapshot(self, api_client, settings, municipio):
        settings.POWER_BI_SERVICE_TOKEN = "power-bi-test-token"
        cache.clear()
        meta = WorkPlanMetaFactory(numero=1)
        create_action(meta, "1.1", municipio)
        first_snapshot = export_to_power_bi()
        create_action(meta, "1.2", municipio)
        second_snapshot = export_to_power_bi()

        response = api_client.get(
            POWER_BI_URL,
            HTTP_AUTHORIZATION="Token power-bi-test-token",
        )

        assert second_snapshot["atualizado_em"] >= first_snapshot["atualizado_em"]
        assert response.data["atualizado_em"] == second_snapshot["atualizado_em"]
        assert len(response.data["resultados"]) == 2

    def test_cached_snapshot_reads_under_three_hundred_milliseconds(
        self, api_client, settings, municipio
    ):
        settings.POWER_BI_SERVICE_TOKEN = "power-bi-test-token"
        cache.clear()
        meta = WorkPlanMetaFactory(numero=1)
        for index in range(1, 101):
            create_action(meta, f"1.{index}", municipio)
        export_to_power_bi()

        started_at = monotonic()
        response = api_client.get(
            POWER_BI_URL,
            HTTP_AUTHORIZATION="Token power-bi-test-token",
        )

        assert response.status_code == 200
        assert monotonic() - started_at < 0.3

    def test_hourly_snapshot_task_is_scheduled(self, settings):
        schedule = settings.CELERY_BEAT_SCHEDULE["export_to_power_bi"]

        assert schedule["task"] == "sgp.tasks.export_to_power_bi"
        assert schedule["schedule"].minute == {0}
