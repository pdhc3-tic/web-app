from __future__ import annotations

import json
import os
from collections import Counter
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError

from apps.core.seed_package import (
    ENTITY_FILES,
    PACKAGE_VERSION,
    decimal,
    find_workbook,
    integer,
    package_record,
    parse_idh,
    read_sheet,
    row_value,
    sha256_file,
    source_id,
    text,
    write_jsonl,
    normalized,
)


class Command(BaseCommand):
    help = "Converte XLSX legados em um pacote normalizado para o seed de produção."

    CULTURE_CATEGORIES = {
        "fruticultura": "frutas",
        "horticultura": "hortalicas",
        "forragem": "forrageiras",
        "plantas medicinais": "medicinais",
    }
    ANIMAL_CATEGORIES = {
        "bovino": "bovino",
        "suinocultura": "suino",
        "caprinocultura": "caprino",
        "ovinocultura": "ovino",
        "apicultura": "apicultura",
        "meliponea": "apicultura",
        "ave caipira": "aves",
        "avicultura": "aves",
        "guine": "aves",
    }
    DIRECT_SGP_SHEETS = {"Pessoas"}
    REDUNDANT_SGP_SHEETS = {"Estados"}

    def add_arguments(self, parser):
        parser.add_argument(
            "--source-dir",
            default=os.getenv("SEED_DATA_DIR", ""),
            help="Diretório dos XLSX legados (ou use SEED_DATA_DIR).",
        )
        parser.add_argument(
            "--output-dir",
            default=os.getenv("SEED_PACKAGE_DIR", ""),
            help="Diretório do pacote (ou use SEED_PACKAGE_DIR).",
        )

    def handle(self, *args, **options):
        source_dir = Path(options["source_dir"]).expanduser()
        output_dir = Path(options["output_dir"]).expanduser()
        if not options["source_dir"]:
            raise CommandError("Informe --source-dir ou configure SEED_DATA_DIR.")
        if not options["output_dir"]:
            raise CommandError("Informe --output-dir ou configure SEED_PACKAGE_DIR.")
        if not source_dir.is_dir():
            raise CommandError(f"Diretório de seed não encontrado: {source_dir}")
        if output_dir.resolve() == source_dir.resolve():
            raise CommandError("O pacote não pode ser gravado no diretório dos XLSX.")

        output_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
        output_dir.chmod(0o700)
        self.counts = Counter()
        self.issues: list[dict[str, Any]] = []
        self.records = {entity: [] for entity in ENTITY_FILES}

        try:
            self.prepare_reference_data(source_dir)
            self.prepare_people(source_dir / "SGP.xlsx")
            self.prepare_sgp_report(source_dir / "SGP.xlsx")
        except ValueError as exc:
            raise CommandError(str(exc)) from exc

        files = {}
        for entity, filename in ENTITY_FILES.items():
            path = output_dir / filename
            digest = write_jsonl(path, self.records[entity])
            path.chmod(0o600)
            files[filename] = {
                "entity": entity,
                "records": len(self.records[entity]),
                "sha256": digest,
            }

        manifest = {
            "package_version": PACKAGE_VERSION,
            "inputs": [
                {
                    "file": path.name,
                    "sha256": sha256_file(path),
                }
                for path in sorted(source_dir.glob("*.xlsx"))
            ],
            "files": files,
            "report": "report.json",
        }
        (output_dir / "manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        (output_dir / "manifest.json").chmod(0o600)
        report = {
            "package_version": PACKAGE_VERSION,
            "counts": dict(sorted(self.counts.items())),
            "issues": self.issues,
        }
        (output_dir / "report.json").write_text(
            json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        (output_dir / "report.json").chmod(0o600)

        self.stdout.write(self.style.SUCCESS(f"Pacote preparado em {output_dir}."))
        for key, value in sorted(self.counts.items()):
            self.stdout.write(f"  {key}: {value}")
        if self.issues:
            self.stdout.write(
                self.style.WARNING(
                    f"  issues: {len(self.issues)} (detalhes em report.json)"
                )
            )

    def prepare_reference_data(self, source_dir: Path):
        municipios_file = find_workbook(source_dir, "municipios")
        comunidades_file = find_workbook(source_dir, "comunidades")
        catalogos_file = find_workbook(source_dir, "culturaseanimais")

        self.prepare_states(municipios_file)
        territories = self.prepare_territories(comunidades_file)
        municipality_data = self.read_municipality_data(municipios_file)
        municipalities = self.prepare_municipalities(
            municipios_file, territories, municipality_data
        )
        self.prepare_communities(comunidades_file, municipalities)
        self.prepare_catalogs(catalogos_file)

    def prepare_states(self, workbook_path: Path):
        for row_number, row in read_sheet(workbook_path, "Estados"):
            sigla = text(row_value(row, "Estado")).upper()
            nome = text(row_value(row, "Nome")) or sigla
            legacy_id = sigla
            if not sigla or len(sigla) != 2:
                self.add_issue("states", workbook_path, "Estados", row_number, legacy_id, "UF inválida")
                continue
            self.records["states"].append(
                package_record(
                    entity="states",
                    file_name=workbook_path.name,
                    sheet="Estados",
                    row=row_number,
                    legacy_id=legacy_id,
                    data={"sigla": sigla, "nome": nome},
                )
            )
            self.counts["states.ready"] += 1

    def prepare_territories(self, workbook_path: Path) -> dict[str, dict[str, Any]]:
        result = {}
        for row_number, row in read_sheet(workbook_path, "Territórios"):
            legacy_id = source_id(row_value(row, "Row ID"))
            name = text(row_value(row, "Território"))
            states = sorted(
                {
                    state.strip().upper()
                    for state in text(row_value(row, "Estados")).split(",")
                    if len(state.strip()) == 2
                }
            )
            if not legacy_id or not name:
                self.add_issue("territories", workbook_path, "Territórios", row_number, legacy_id, "ID ou nome ausente")
                continue
            data = {"nome": name, "estados": states, "ativo": True}
            result[legacy_id] = data
            self.records["territories"].append(
                package_record(
                    entity="territories",
                    file_name=workbook_path.name,
                    sheet="Territórios",
                    row=row_number,
                    legacy_id=legacy_id,
                    data=data,
                )
            )
            self.counts["territories.ready"] += 1
        return result

    def read_municipality_data(self, workbook_path: Path) -> dict[str, dict[str, Any]]:
        result = {}
        for _, row in read_sheet(workbook_path, "Municipios_Dados"):
            code = source_id(row_value(row, "Row ID", "Município"))
            if code:
                result[code] = row
        return result

    def prepare_municipalities(self, workbook_path, territories, municipality_data):
        result = {}
        for row_number, row in read_sheet(workbook_path, "Municípios"):
            code = source_id(row_value(row, "Row ID"))
            name = text(row_value(row, "Município"))
            state = text(row_value(row, "Estado")).upper()
            if not code or not name or len(state) != 2:
                self.add_issue("municipalities", workbook_path, "Municípios", row_number, code, "Código, nome ou UF inválido")
                continue
            territory_id = source_id(row_value(row, "Território"))
            data = {
                "codigo_ibge": code,
                "nome": name,
                "sigla_estado": state,
                "territorio_id": territory_id,
            }
            extra = municipality_data.get(code)
            if extra:
                data.update(
                    {
                        "area_km2": decimal(row_value(extra, "Área")),
                        "pop_total": integer(row_value(extra, "População Total")),
                        "pop_rural": integer(row_value(extra, "População Rural")),
                        "idh": parse_idh(row_value(extra, "IDH")),
                        "perc_extr_pobres": decimal(row_value(extra, "% de Extremamente Pobres")),
                        "benef_programa_agri_familiar": integer(row_value(extra, "Beneficiários do Agricultura Familiar")),
                        "estab_agri_familiar": integer(row_value(extra, "Estabelecimentos Agricultura Familiar")),
                    }
                )
            quality = []
            if territory_id and territory_id not in territories:
                quality.append("unmapped_territory")
                self.add_issue("municipalities", workbook_path, "Municípios", row_number, code, f"Território inexistente: {territory_id}")
            result[code] = data
            self.records["municipalities"].append(
                package_record(
                    entity="municipalities",
                    file_name=workbook_path.name,
                    sheet="Municípios",
                    row=row_number,
                    legacy_id=code,
                    data=data,
                    quality=quality,
                )
            )
            self.counts["municipalities.ready"] += 1
        return result

    def prepare_communities(self, workbook_path, municipalities):
        for row_number, row in read_sheet(workbook_path, "Comunidades"):
            legacy_id = source_id(row_value(row, "Row ID"))
            name = text(row_value(row, "Comunidade"))
            municipality_code = source_id(row_value(row, "Município"))
            if not legacy_id or not name:
                self.add_issue("communities", workbook_path, "Comunidades", row_number, legacy_id, "ID ou nome ausente")
                continue
            if municipality_code not in municipalities:
                self.add_issue("communities", workbook_path, "Comunidades", row_number, legacy_id, f"Município inexistente: {municipality_code}")
                continue
            coordinates = text(row_value(row, "Localização Geográfica")).split(",")
            data = {
                "nome": name,
                "municipality_code": municipality_code,
                "lat": decimal(coordinates[0], "0.000001") if len(coordinates) == 2 else None,
                "lng": decimal(coordinates[1], "0.000001") if len(coordinates) == 2 else None,
            }
            self.records["communities"].append(
                package_record(
                    entity="communities",
                    file_name=workbook_path.name,
                    sheet="Comunidades",
                    row=row_number,
                    legacy_id=legacy_id,
                    data=data,
                )
            )
            self.counts["communities.ready"] += 1

    def prepare_catalogs(self, workbook_path):
        for row_number, row in read_sheet(workbook_path, "Culturas"):
            legacy_id = source_id(row_value(row, "Row ID"))
            name = text(row_value(row, "Cultura"))
            legacy_category = normalized(row_value(row, "Tipo"))
            if not legacy_id or not name:
                self.add_issue("cultures", workbook_path, "Culturas", row_number, legacy_id, "ID ou nome ausente")
                continue
            category = self.CULTURE_CATEGORIES.get(legacy_category, "outras")
            quality = []
            if category == "outras" and legacy_category not in ("", "outras"):
                quality.append("default_category")
                self.counts["cultures.default_category"] += 1
            self.records["cultures"].append(
                package_record(
                    entity="cultures",
                    file_name=workbook_path.name,
                    sheet="Culturas",
                    row=row_number,
                    legacy_id=legacy_id,
                    data={"nome": name, "categoria": category, "ciclo": "anual"},
                    quality=quality,
                )
            )
            self.counts["cultures.ready"] += 1

        for row_number, row in read_sheet(workbook_path, "Pecuária"):
            legacy_id = source_id(row_value(row, "Row ID"))
            name = text(row_value(row, "Pecuária", "Espécie"))
            if not legacy_id or not name:
                self.add_issue("animals", workbook_path, "Pecuária", row_number, legacy_id, "ID ou nome ausente")
                continue
            category = self.ANIMAL_CATEGORIES.get(normalized(name), "outros")
            quality = ["default_category"] if category == "outros" else []
            if quality:
                self.counts["animals.default_category"] += 1
            self.records["animals"].append(
                package_record(
                    entity="animals",
                    file_name=workbook_path.name,
                    sheet="Pecuária",
                    row=row_number,
                    legacy_id=legacy_id,
                    data={"nome": name, "categoria": category},
                    quality=quality,
                )
            )
            self.counts["animals.ready"] += 1
        self.counts["seeds.not_supported"] += 1
        self.issues.append(
            {
                "type": "unsupported_sheet",
                "file": workbook_path.name,
                "sheet": "Sementes Crioulas",
                "reason": "não existe modelo correspondente",
            }
        )

    def prepare_people(self, workbook_path: Path):
        if not workbook_path.exists():
            self.issues.append(
                {
                    "type": "missing_file",
                    "file": workbook_path.name,
                    "reason": "arquivo opcional não encontrado",
                }
            )
            return
        seen = {}
        for row_number, row in read_sheet(workbook_path, "Pessoas"):
            legacy_id = source_id(row_value(row, "ID"))
            name = text(row_value(row, "Nome"))
            email = text(row_value(row, "Email")).lower()
            if not legacy_id or not name or "@" not in email:
                self.add_issue("people", workbook_path, "Pessoas", row_number, legacy_id, "ID, nome ou e-mail inválido")
                continue
            if email in seen:
                self.counts["people.duplicate_email"] += 1
                self.issues.append(
                    {
                        "type": "duplicate",
                        "entity": "people",
                        "file": workbook_path.name,
                        "sheet": "Pessoas",
                        "row": row_number,
                        "id": legacy_id,
                        "reason": f"e-mail já usado pela linha {seen[email]}",
                    }
                )
                continue
            seen[email] = row_number
            self.records["people"].append(
                package_record(
                    entity="people",
                    file_name=workbook_path.name,
                    sheet="Pessoas",
                    row=row_number,
                    legacy_id=legacy_id,
                    data={
                        "email": email,
                        "nome": name,
                        "telefone": source_id(row_value(row, "Telefone")),
                    },
                )
            )
            self.counts["people.ready"] += 1

    def prepare_sgp_report(self, workbook_path: Path):
        if not workbook_path.exists():
            return
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for sheet in workbook.sheetnames:
                if sheet in self.DIRECT_SGP_SHEETS:
                    continue
                if sheet in self.REDUNDANT_SGP_SHEETS:
                    self.issues.append(
                        {
                            "type": "redundant_sheet",
                            "file": workbook_path.name,
                            "sheet": sheet,
                            "reason": "dados já vêm de outro XLSX",
                        }
                    )
                    continue
                self.counts["sgp.sheets_not_supported"] += 1
                self.issues.append(
                    {
                        "type": "unsupported_sheet",
                        "file": workbook_path.name,
                        "sheet": sheet,
                        "reason": "não existe mapeamento direto para o schema atual",
                    }
                )
        finally:
            workbook.close()

    def add_issue(self, entity, workbook_path, sheet, row, legacy_id, reason):
        self.counts[f"{entity}.invalid"] += 1
        self.issues.append(
            {
                "type": "invalid",
                "entity": entity,
                "file": workbook_path.name,
                "sheet": sheet,
                "row": row,
                "id": legacy_id,
                "reason": reason,
            }
        )
