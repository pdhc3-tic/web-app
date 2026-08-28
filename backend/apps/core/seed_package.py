from __future__ import annotations

import hashlib
import json
import unicodedata
from collections.abc import Iterable
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PACKAGE_VERSION = 1
ENTITY_FILES = {
    "states": "states.jsonl",
    "territories": "territories.jsonl",
    "municipalities": "municipalities.jsonl",
    "communities": "communities.jsonl",
    "cultures": "cultures.jsonl",
    "animals": "animals.jsonl",
    "people": "people.jsonl",
}


def normalized(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(
        char for char in text if not unicodedata.combining(char)
    ).strip().lower()


def header_key(value: Any) -> str:
    return normalized(value).replace(" ", "").replace("_", "")


def source_id(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def integer(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return None


def decimal(value: Any, places: str = "0.01") -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(
            Decimal(places), rounding=ROUND_HALF_UP
        )
    except (InvalidOperation, ValueError):
        return None


def parse_idh(value: Any) -> Decimal | None:
    parsed = decimal(value, "0.001")
    if parsed is not None and parsed > 1:
        parsed = (parsed / Decimal("10000")).quantize(
            Decimal("0.001"), rounding=ROUND_HALF_UP
        )
    return parsed


def canonical_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    return value


def read_sheet(workbook_path: Path, sheet_name: str) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"A planilha '{sheet_name}' não existe em {workbook_path.name}."
            )
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        raw_header = next(rows, None)
        if raw_header is None:
            return []
        headers = [header_key(value) for value in raw_header]
        result = []
        for row_number, raw_row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in raw_row):
                continue
            result.append(
                (
                    row_number,
                    {
                        header: raw_row[index] if index < len(raw_row) else None
                        for index, header in enumerate(headers)
                        if header
                    },
                )
            )
        return result
    finally:
        workbook.close()


def row_value(row: dict[str, Any], *names: str) -> Any:
    for name in names:
        value = row.get(header_key(name))
        if value not in (None, ""):
            return value
    return None


def find_workbook(source_dir: Path, *prefixes: str) -> Path:
    normalized_prefixes = tuple(
        normalized(prefix).replace(" ", "") for prefix in prefixes
    )
    matches = [
        path
        for path in source_dir.glob("*.xlsx")
        if any(
            normalized(path.stem).replace(" ", "").startswith(prefix)
            for prefix in normalized_prefixes
        )
    ]
    if not matches:
        raise ValueError(
            f"Nenhum XLSX encontrado para {', '.join(prefixes)} em {source_dir}."
        )
    if len(matches) > 1:
        names = ", ".join(path.name for path in sorted(matches))
        raise ValueError(f"Mais de um XLSX corresponde a {prefixes}: {names}")
    return matches[0]


def package_record(
    *,
    entity: str,
    file_name: str,
    sheet: str,
    row: int,
    legacy_id: str,
    data: dict[str, Any],
    quality: Iterable[str] = (),
) -> dict[str, Any]:
    return {
        "source": {
            "file": file_name,
            "sheet": sheet,
            "row": row,
            "id": legacy_id,
        },
        "entity": entity,
        "quality": sorted(set(quality)),
        "data": {
            key: canonical_value(value) for key, value in sorted(data.items())
        },
    }


def write_jsonl(path: Path, records: list[dict[str, Any]]) -> str:
    ordered = sorted(
        records,
        key=lambda record: (
            record["source"]["sheet"],
            record["source"]["id"],
            record["source"]["row"],
        ),
    )
    content = "".join(
        json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n"
        for record in ordered
    )
    path.write_text(content, encoding="utf-8")
    return hashlib.sha256(content.encode("utf-8")).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for block in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()
